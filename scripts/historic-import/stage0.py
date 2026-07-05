#!/usr/bin/env python3
"""PP6 historic-import — Stage 0 (staging, READ-ONLY).

Parses one season's weekly workbooks (PP6 <season> Week N.xlsx/.xlsm) plus the
LeagueSafe payment CSV into reviewable JSON under data/imports/historic/_staging/<season>/.

Emits:
  games.json          - the weekly game slate (favorite/dog/spread/scores)
  picks.json          - every player's picks, lock, and re-derived score
  payments.json       - LeagueSafe payment rows
  team_names.json     - distinct team names seen (to build the CFBD/DB alias table)
  players.json        - distinct players (name + emails) for identity matching
  reconciliation.json - re-derived weekly points vs the workbook's own totals
  summary.json        - counts + any parse warnings

NO database writes. Usage:
  python3 scripts/historic-import/stage0.py 2016
  python3 scripts/historic-import/stage0.py 2016 2017 2018
"""
import glob
import json
import os
import re
import sys

import openpyxl

from common import (
    ARCHIVE_DIR, STAGING_DIR, parse_pick_string, norm_team, norm_email,
    player_key, full_name, score_pick, team_slug, game_slug,
)

# Optional " - " between the year and Master/Week (2023 naming), case-insensitive.
MASTER_FILE_RE = re.compile(r"PP6 (\d{4})\s*-?\s*Master\b.*\.xls[mx]$", re.IGNORECASE)
WEEK_FILE_RE = re.compile(r"PP6 (\d{4})\s*-?\s*Week (\d+)\b(.*)\.xls[mx]$", re.IGNORECASE)


def find_source_workbook(season_dir, season):
    """Return the single cumulative source workbook for a season.

    Both Master files and weekly files are cumulative. Candidates are every
    Master plus the highest-numbered weekly file; we pick the most recently
    modified (a late-season weekly can be newer/more complete than an early
    'MASTER v2'). Excel lock/temp files (~$...) are ignored.
    """
    masters = []
    weeks = {}
    for path in glob.glob(os.path.join(season_dir, "*.xls*")):
        base = os.path.basename(path)
        if base.startswith("~$"):
            continue
        m = MASTER_FILE_RE.match(base)
        if m and int(m.group(1)) == season:
            masters.append((path, os.path.getmtime(path), "master"))
            continue
        w = WEEK_FILE_RE.match(base)
        if w and int(w.group(1)) == season:
            wk = int(w.group(2))
            mt = os.path.getmtime(path)
            if wk not in weeks or mt > weeks[wk][1]:
                weeks[wk] = (path, mt)
    candidates = list(masters)
    if weeks:
        last = max(weeks)
        candidates.append((weeks[last][0], weeks[last][1], f"week{last}"))
    if not candidates:
        return None, None
    candidates.sort(key=lambda x: x[1], reverse=True)  # newest first
    return candidates[0][0], candidates[0][2]


def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def _week_num(cell):
    if cell is None:
        return None
    m = re.search(r"(\d+)", str(cell))
    return int(m.group(1)) if m else None


GAME_CODE_RE = re.compile(r"W(\d+)G(\d+)", re.IGNORECASE)


def parse_games_sheet(ws, season, warnings):
    """Parse the 'Games' sheet of a cumulative workbook -> list of game dicts,
    one per (week, slot).

    Robust across all archive layouts (2016-2024):
      - week/slot come from the col0 code 'W<week>G<slot>' (2019 stores slot as
        a bare int, so we don't rely on col2).
      - col4 'Dog' team, col5 'Favorite' team, col6 favorite spread (negative)
        are stable everywhere.
      - favorite/dog SCORES are the *second* 'Favorite'/'Dog' header columns
        (col7/8 in 2016/17/20+, col10/11 in 2018/19).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append("Games sheet empty")
        return []
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]

    fav_hdrs = [i for i, h in enumerate(hdr) if h == "Favorite"]
    dog_hdrs = [i for i, h in enumerate(hdr) if h == "Dog"]
    if len(fav_hdrs) < 2 or len(dog_hdrs) < 2:
        warnings.append(f"Games sheet: could not locate score columns "
                        f"(Favorite hdrs={fav_hdrs}, Dog hdrs={dog_hdrs})")
        return []
    c_fav_team, c_fav_score = fav_hdrs[0], fav_hdrs[1]
    c_dog_team, c_dog_score = dog_hdrs[0], dog_hdrs[1]
    c_spread = 6  # favorite spread, stable across layouts

    games = []
    for row in rows[1:]:
        m = GAME_CODE_RE.match(str(_cell(row, 0) or "").strip())
        if not m:
            continue
        week, slot = int(m.group(1)), int(m.group(2))
        dog = _cell(row, c_dog_team)
        fav = _cell(row, c_fav_team)
        if fav is None or dog is None:
            continue
        fav_spread = _cell(row, c_spread)
        try:
            fav_spread = float(fav_spread) if fav_spread is not None else None
        except (TypeError, ValueError):
            fav_spread = None
        fav_score = _cell(row, c_fav_score)
        dog_score = _cell(row, c_dog_score)
        completed = isinstance(fav_score, (int, float)) and isinstance(dog_score, (int, float))
        mag = abs(fav_spread) if fav_spread is not None else None
        games.append({
            "season": season, "week": week, "slot": slot,
            "favorite": str(fav).strip(), "underdog": str(dog).strip(),
            "favorite_spread": fav_spread,          # negative
            "favorite_score": fav_score if completed else None,
            "underdog_score": dog_score if completed else None,
            "fav_slug": game_slug(str(fav), mag) if mag is not None else None,
            "dog_slug": game_slug(str(dog), mag) if mag is not None else None,
            "status": "completed" if completed else "scheduled",
        })
    if not games:
        warnings.append("no games parsed from Games sheet")
    return games


def parse_picks_sheet(ws, season, games_by_week_slot, warnings):
    """Parse the 'Picks' sheet -> list of pick dicts (one per selected game).

    Uses header names to locate columns (robust to minor reordering):
      First/Last Name, Form/Logged-in Email, Game 1..15 (pick strings),
      'Which game is your LOCK?', and the workbook's own 'Points' total.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append("Picks sheet empty")
        return [], []
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]

    def find(label, start=0):
        for i in range(start, len(hdr)):
            if hdr[i] == label:
                return i
        return None

    def find_contains(substr, start=0):
        for i in range(start, len(hdr)):
            if substr.lower() in hdr[i].lower():
                return i
        return None

    c_first = find("First Name")
    c_last = find("Last Name")
    c_form_email = find_contains("Form Email") or find_contains("Email")
    c_login_email = find_contains("Logged-in Email")
    c_status = next((i for i in range(len(hdr)) if hdr[i].lower() == "status"), None)
    c_week = find("Week")
    c_lock = find_contains("LOCK")
    # first block of "Game N:" columns holds the pick strings; a second block
    # of "Game N" columns (no colon) holds the workbook's own per-pick points
    # (loss = -1 marker, push = 10, win = actual points incl. that year's rules).
    game_cols = {}
    pts_cols = {}
    for i, h in enumerate(hdr):
        m = re.match(r"^Game (\d+)\s*:", h)
        if m:
            game_cols[int(m.group(1))] = i
            continue
        m2 = re.match(r"^Game (\d+)\s*$", h)
        if m2:
            pts_cols[int(m2.group(1))] = i
    # the workbook's own computed weekly points (second 'Points' header)
    c_points = None
    for i in range(len(hdr) - 1, -1, -1):
        if hdr[i] == "Points":
            c_points = i
            break

    if not game_cols:
        warnings.append("Picks sheet - no 'Game N:' pick columns found")
        return [], []
    if c_week is None:
        warnings.append("Picks sheet - no 'Week' column found")
        return [], []

    picks = []
    recon = []  # per player-week: re-derived total vs workbook total
    for row in rows[1:]:
        first = _cell(row, c_first) if c_first is not None else None
        last = _cell(row, c_last) if c_last is not None else None
        form_email = _cell(row, c_form_email) if c_form_email is not None else None
        login_email = _cell(row, c_login_email) if c_login_email is not None else None
        if not (first or last or form_email or login_email):
            continue
        week = _week_num(_cell(row, c_week))
        if week is None:
            continue
        status = str(_cell(row, c_status)).strip() if c_status is not None and _cell(row, c_status) else ""
        # Skip superseded resubmissions (2024+ marks the stale copies 'EXCLUDE';
        # the surviving row is 'LAST'/'OK'). Prevents multi-counting one player.
        if status.upper() == "EXCLUDE":
            continue
        # Prefer the FORM email (the identity the player entered for this entry)
        # over the logged-in email: families/multi-entry often submit under one
        # shared login, so the form email is what distinguishes the actual player.
        pkey = player_key(first, last, form_email, login_email)
        name = full_name(first, last)

        lock_slot = None
        if c_lock is not None:
            lv = _cell(row, c_lock)
            if lv:
                mm = re.search(r"(\d+)", str(lv))
                if mm:
                    lock_slot = int(mm.group(1))

        player_picks = []
        derived_total = 0
        for slot, col in game_cols.items():
            raw = _cell(row, col)
            if raw is None or not str(raw).strip():
                continue
            raw = str(raw).strip()
            game = games_by_week_slot.get((week, slot))
            is_lock = (slot == lock_slot)

            # determine which side the player picked (handles both encodings:
            # readable "Team +/-spread" and 2019+ slugs like "north-carolina-23-0")
            side = None
            if game:
                parsed = parse_pick_string(raw)
                if parsed:                              # readable string
                    pteam = norm_team(parsed[0])
                    if pteam == norm_team(game["favorite"]):
                        side = "favorite"
                    elif pteam == norm_team(game["underdog"]):
                        side = "underdog"
                else:                                    # slug form
                    key = raw.lower()
                    if key == (game.get("fav_slug") or "") or key.startswith(team_slug(game["favorite"]) + "-"):
                        side = "favorite"
                    elif key == (game.get("dog_slug") or "") or key.startswith(team_slug(game["underdog"]) + "-"):
                        side = "underdog"

            if side == "favorite":
                sel_team, picked_spread = game["favorite"], game["favorite_spread"]
                picked_score, opp_score = game["favorite_score"], game["underdog_score"]
            elif side == "underdog":
                sel_team = game["underdog"]
                picked_spread = -game["favorite_spread"] if game["favorite_spread"] is not None else None
                picked_score, opp_score = game["underdog_score"], game["favorite_score"]
            else:
                sel_team = picked_spread = picked_score = opp_score = None

            # authoritative points/result from the workbook's own per-pick block
            wb_val = _cell(row, pts_cols[slot]) if slot in pts_cols else None
            wb_val = wb_val if isinstance(wb_val, (int, float)) else None
            if wb_val is None:
                result_wb, points_wb = None, None
            elif wb_val <= 0:
                result_wb, points_wb = "loss", 0        # -1 (or 0) = loss marker
            elif wb_val == 10:
                result_wb, points_wb = "push", 10
            else:
                result_wb, points_wb = "win", int(wb_val)

            rec = {
                "season": season, "week": week, "slot": slot,
                "player_key": pkey, "player_name": name,
                "email_form": norm_email(form_email),
                "email_login": norm_email(login_email),
                "status": status,
                "raw_pick": raw,
                "selected_team": sel_team, "selected_spread": picked_spread,
                "side": side,
                "is_lock": is_lock,
                # authoritative (workbook's own scoring, preserves that year's rules):
                "result": result_wb, "points_earned": points_wb,
                # cross-check (re-derived with current DB rules):
                "result_derived": None, "points_derived": None,
                "matched_game": bool(game),
            }
            if not game:
                rec["match_warning"] = "no game for (week, slot)"
            elif side is None:
                rec["match_warning"] = "pick did not match favorite or underdog"
            elif game["status"] == "completed" and picked_score is not None:
                result, pts = score_pick(picked_spread, picked_score, opp_score, is_lock)
                rec["result_derived"], rec["points_derived"] = result, pts
                derived_total += pts
                if points_wb is not None and pts != points_wb:
                    rec["scoring_diff"] = {"derived": pts, "workbook": points_wb}
            player_picks.append(rec)

        if not player_picks:
            continue

        wb_total = _cell(row, c_points) if c_points is not None else None
        wb_total = wb_total if isinstance(wb_total, (int, float)) else None
        block_sum = sum(p["points_earned"] or 0 for p in player_picks)
        # In lock-doubling years the official 'Points' total exceeds the per-pick
        # block sum by the lock's doubled bonus; attribute that residual to the
        # lock pick so per-pick points_earned sum to the official leaderboard total.
        if wb_total is not None and abs(block_sum - wb_total) > 0.001:
            delta = wb_total - block_sum
            lock_pick = next((p for p in player_picks if p["is_lock"] and p["points_earned"]), None)
            if lock_pick is not None:
                lock_pick["points_earned"] += delta
                lock_pick["lock_bonus_adjustment"] = delta
            else:
                for p in player_picks:
                    p["points_total_mismatch"] = True  # flag for review at load
        picks.extend(player_picks)

        wb_pick_sum = sum(p["points_earned"] or 0 for p in player_picks)
        recon.append({
            "season": season, "week": week, "player_key": pkey, "player_name": name,
            "status": status,
            "num_picks": len(player_picks),
            "lock_slot": lock_slot,
            "points_derived": derived_total,       # re-derived (current rules) cross-check
            "points_workbook": wb_total,           # workbook's own 'Points' total (authoritative)
            "points_workbook_pick_sum": wb_pick_sum,  # sum of per-pick block (integrity check)
            "match": (wb_total is not None and abs(derived_total - wb_total) < 0.001),
        })
    return picks, recon


def parse_payments(season_dir, season, warnings):
    """Parse LeagueSafe payments for a season -> (rows, source_filename).

    Handles both formats seen in the archive:
      - 'Pigskin Pick Six payment details*.csv' (Owner/OwnerEmail/EntryFee/Paid/Status/IsCommish)
      - 'Leaguesafe.xlsx' (Owner/First/Last/Email/Entry Fee/Paid/Pending/Owes)
    """
    import csv as _csv
    csvs = [p for p in glob.glob(os.path.join(season_dir, "*.csv"))
            if "payment" in os.path.basename(p).lower()]
    if csvs:
        csvs.sort(key=lambda p: (str(season) in os.path.basename(p), os.path.getmtime(p)), reverse=True)
        path = csvs[0]
        out = []
        import io
        raw = open(path, "rb").read()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp1252")          # LeagueSafe exports use Windows-1252
        with io.StringIO(text) as f:
            for r in _csv.DictReader(f):
                owner = (r.get("Owner") or "").strip()
                email = norm_email(r.get("OwnerEmail"))
                if not owner and not email:
                    continue
                out.append({
                    "season": season, "owner": owner, "email": email,
                    "entry_fee": r.get("EntryFee"), "paid": r.get("Paid"),
                    "pending": r.get("Pending"), "owes": r.get("Owes"),
                    "status": (r.get("Status") or "").strip(),
                    "is_commish": (r.get("IsCommish") or "").strip(),
                })
        return out, os.path.basename(path)

    xlsxs = [p for p in glob.glob(os.path.join(season_dir, "*.xls*"))
             if "leaguesafe" in os.path.basename(p).lower()]
    if xlsxs:
        path = sorted(xlsxs, key=os.path.getmtime, reverse=True)[0]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], os.path.basename(path)
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        col = {h: i for i, h in enumerate(hdr)}
        out = []
        for r in rows[1:]:
            owner = str(_cell(r, col.get("Owner"))).strip() if _cell(r, col.get("Owner")) else ""
            email = norm_email(_cell(r, col.get("Email")))
            if not owner and not email:
                continue
            out.append({
                "season": season, "owner": owner, "email": email,
                "entry_fee": _cell(r, col.get("Entry Fee")), "paid": _cell(r, col.get("Paid")),
                "pending": _cell(r, col.get("Pending")), "owes": _cell(r, col.get("Owes")),
                "status": "", "is_commish": "",
            })
        return out, os.path.basename(path)

    warnings.append(f"{season}: no payment file (CSV or Leaguesafe.xlsx) found")
    return [], None


def get_sheet(wb, *names):
    lower = {s.lower(): s for s in wb.sheetnames}
    for n in names:
        if n.lower() in lower:
            return wb[lower[n.lower()]]
    return None


def process_season(season):
    season_dir = os.path.join(ARCHIVE_DIR, str(season))
    if not os.path.isdir(season_dir):
        print(f"  !! season dir not found: {season_dir}")
        return
    warnings = []
    src_path, src_kind = find_source_workbook(season_dir, season)
    if not src_path:
        print(f"  !! no Master or weekly workbook found for {season}")
        return
    print(f"  source workbook ({src_kind}): {os.path.basename(src_path)}")

    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    games_ws = get_sheet(wb, "Games")
    picks_ws = get_sheet(wb, "Picks")
    all_games = parse_games_sheet(games_ws, season, warnings) if games_ws else []
    if games_ws is None:
        warnings.append("no 'Games' sheet in source workbook")
    by_week_slot = {(g["week"], g["slot"]): g for g in all_games}
    if picks_ws is None:
        warnings.append("no 'Picks' sheet in source workbook")
        all_picks, all_recon = [], []
    else:
        all_picks, all_recon = parse_picks_sheet(picks_ws, season, by_week_slot, warnings)
    wb.close()

    # per-week console summary
    weeks = sorted({g["week"] for g in all_games})
    for w in weeks:
        wr = [r for r in all_recon if r["week"] == w]
        print(f"    W{w}: {sum(1 for g in all_games if g['week']==w)} games, "
              f"{sum(1 for p in all_picks if p['week']==w)} picks, "
              f"{sum(1 for r in wr if r['match'])}/{len(wr)} score-matched")

    payments, pay_file = parse_payments(season_dir, season, warnings)

    # distinct teams (to build the CFBD/DB alias table later)
    teams = sorted({g["favorite"] for g in all_games} | {g["underdog"] for g in all_games})
    # distinct players
    players = {}
    for p in all_picks:
        k = p["player_key"]
        if k not in players:
            players[k] = {"player_key": k, "name": p["player_name"],
                          "email_form": p["email_form"], "email_login": p["email_login"]}

    # write outputs
    out_dir = os.path.join(STAGING_DIR, str(season))
    os.makedirs(out_dir, exist_ok=True)

    def dump(name, obj):
        with open(os.path.join(out_dir, name), "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2, default=str)

    matched = sum(1 for r in all_recon if r["match"])
    # integrity: does the per-pick workbook block sum to the workbook 'Points' total?
    integrity_ok = sum(1 for r in all_recon if r["points_workbook"] is not None
                       and abs(r["points_workbook_pick_sum"] - r["points_workbook"]) < 0.001)
    scoring_diffs = sum(1 for p in all_picks if "scoring_diff" in p)
    dump("games.json", all_games)
    dump("picks.json", all_picks)
    dump("payments.json", payments)
    dump("team_names.json", teams)
    dump("players.json", list(players.values()))
    dump("reconciliation.json", all_recon)
    summary = {
        "season": season,
        "source_workbook": os.path.basename(src_path),
        "weeks": weeks,
        "payment_source": pay_file,
        "counts": {
            "games": len(all_games),
            "picks": len(all_picks),
            "players": len(players),
            "payments": len(payments),
            "distinct_teams": len(teams),
        },
        "workbook_points_integrity": {
            "player_weeks_checked": sum(1 for r in all_recon if r["points_workbook"] is not None),
            "per_pick_sum_equals_total": integrity_ok,
            "picks_where_derived_differs_from_workbook": scoring_diffs,
        },
        "score_reconciliation": {
            "note": "cross-check only: re-derived (current rules) vs workbook total. "
                    "Authoritative points/result come from the workbook per-pick block.",
            "player_weeks": len(all_recon),
            "matched": matched,
            "mismatched": len(all_recon) - matched,
            "match_rate": round(matched / len(all_recon), 4) if all_recon else None,
        },
        "warnings": warnings,
    }
    dump("summary.json", summary)

    print(f"  -> wrote staging to {out_dir}")
    print(f"     games={len(all_games)} picks={len(all_picks)} players={len(players)} "
          f"payments={len(payments)} teams={len(teams)}")
    print(f"     workbook points integrity: {integrity_ok}/"
          f"{summary['workbook_points_integrity']['player_weeks_checked']} player-weeks "
          f"(per-pick sum == total); derived differs on {scoring_diffs} picks (cross-check)")
    if warnings:
        print(f"     WARNINGS ({len(warnings)}):")
        for w in warnings[:15]:
            print(f"       - {w}")
        if len(warnings) > 15:
            print(f"       ... and {len(warnings) - 15} more (see summary.json)")


def main():
    seasons = sys.argv[1:] or ["2016"]
    for s in seasons:
        print(f"\n=== Season {s} ===")
        process_season(int(s))


if __name__ == "__main__":
    main()
